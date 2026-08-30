#!/usr/bin/env python3
"""
Gera o Relatório de Gestão de Obras (HTML), no mesmo padrão visual do
modelo "Relatório D6", a partir do workbook Gestao_Obras_Template.xlsx.

Uso:
    python3 gerar_relatorio.py <planilha.xlsx> <ID_OBRA> [saida.html]

Exemplo:
    python3 gerar_relatorio.py Gestao_Obras_Template.xlsx D6 relatorio_D6.html
"""
import sys
import re
import unicodedata
import datetime
from collections import defaultdict
import openpyxl

MESES = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]


def to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    # tenta texto ISO
    try:
        return datetime.date.fromisoformat(str(v)[:10])
    except Exception:
        return None


def to_float(v):
    try:
        return float(v)
    except Exception:
        return 0.0


def fmt_money(v):
    s = f"{v:,.2f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"


def slug_etapa(nome):
    """Converte o nome de uma etapa num id de âncora HTML seguro (sem acento/espaço)."""
    nome = (nome or "").strip()
    if not nome:
        return "etapa-sem-etapa"
    sem_acento = unicodedata.normalize("NFKD", nome).encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", sem_acento).strip("-").lower()
    return f"etapa-{slug}" if slug else "etapa-sem-etapa"


def load_data(path, obra_id):
    wb = openpyxl.load_workbook(path, data_only=False)

    # Obras
    obra_nome = obra_id
    if "Obras" in wb.sheetnames:
        ws = wb["Obras"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            if str(row[0]).strip().upper() == obra_id.strip().upper():
                obra_nome = row[1] or obra_id
                break

    # Etapas -> cor, mantendo ordem do cadastro
    etapas_cor = {}
    ordem_etapas = []
    if "Etapas" in wb.sheetnames:
        ws = wb["Etapas"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            nome = str(row[0]).strip()
            cor = str(row[1]).strip() if row[1] else "999999"
            if not cor.startswith("#"):
                cor = "#" + cor
            etapas_cor[nome] = cor
            ordem_etapas.append(nome)
    # mapa de normalização: compara sem diferenciar maiúscula/minúscula/acento simples,
    # pra "FUNDAÇÃO", "Fundação" e "fundação" caírem na mesma etapa cadastrada
    etapas_lower = {nome.strip().upper(): nome for nome in ordem_etapas}

    def normaliza_etapa(bruto):
        bruto = (bruto or "").strip()
        if not bruto:
            return ""
        return etapas_lower.get(bruto.upper(), bruto)

    # Lançamentos (aba única: pagamento + material juntos)
    pagamentos = []
    materiais = []
    ws = wb["Lançamentos"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or str(row[0]).strip() == "":
            continue
        if str(row[0]).strip().upper() != obra_id.strip().upper():
            continue
        (obra, status, data, venc, fornecedor, etapa, responsavel,
         produto, qtd, custo_und, desconto, _valor_total, tipo_gasto) = (list(row) + [None] * 13)[:13]
        qtd = to_float(qtd)
        custo_und = to_float(custo_und)
        desconto = to_float(desconto)
        valor = qtd * custo_und - desconto
        status_bruto = (status or "").strip().upper()
        if status_bruto == "PAGO":
            status = "Pago"
        elif status_bruto in ("A PAGAR", "APAGAR", "A-PAGAR"):
            status = "A Pagar"
        else:
            status = (status or "").strip()
        etapa = normaliza_etapa(etapa)
        produto = (produto or "").strip()
        if not produto:
            produto = "Verificar"
        if etapa and etapa not in etapas_cor:
            etapas_cor[etapa] = "#999999"
            ordem_etapas.append(etapa)
            etapas_lower[etapa.upper()] = etapa

        pagamentos.append({
            "status": status,
            "data": to_date(data),
            "vencimento": to_date(venc),
            "fornecedor": (fornecedor or "").strip(),
            "etapa": etapa,
            "descricao": produto,
            "valor": valor,
            "responsavel": (responsavel or "").strip(),
            "tipo_gasto": (tipo_gasto or "").strip(),
        })
        materiais.append({
            "produto": produto,
            "etapa": etapa,
            "quantidade": qtd,
            "custo_und": custo_und,
            "custo_total": valor,
        })

    return obra_nome, etapas_cor, ordem_etapas, pagamentos, materiais


def build_report(obra_id, path, hoje=None):
    hoje = hoje or datetime.date.today()
    obra_nome, etapas_cor, ordem_etapas, pagamentos, materiais = load_data(path, obra_id)

    total_pago = sum(p["valor"] for p in pagamentos if p["status"] == "Pago")
    total_a_pagar = sum(p["valor"] for p in pagamentos if p["status"] == "A Pagar")
    total_geral = total_pago + total_a_pagar

    fim_semana = hoje + datetime.timedelta(days=7)
    vencimentos_semana = sum(
        p["valor"] for p in pagamentos
        if p["status"] == "A Pagar" and p["vencimento"] and hoje <= p["vencimento"] <= fim_semana
    )

    # --- Fluxo de caixa mensal ---
    pago_mes = defaultdict(float)
    apagar_mes = defaultdict(float)
    for p in pagamentos:
        if p["status"] == "Pago" and p["data"]:
            key = (p["data"].year, p["data"].month)
            pago_mes[key] += p["valor"]
        elif p["status"] == "A Pagar" and p["vencimento"]:
            key = (p["vencimento"].year, p["vencimento"].month)
            apagar_mes[key] += p["valor"]

    meses_chave = sorted(set(pago_mes) | set(apagar_mes))
    meses_chave = meses_chave[-7:]  # últimos 7 meses com lançamento

    max_val = 0.0
    for k in meses_chave:
        max_val = max(max_val, pago_mes.get(k, 0.0), apagar_mes.get(k, 0.0))
    max_val = max_val or 1.0

    barras_html = []
    for (ano, mes) in meses_chave:
        v_pago = pago_mes.get((ano, mes), 0.0)
        v_apagar = apagar_mes.get((ano, mes), 0.0)
        h_pago = (v_pago / max_val) * 100
        h_apagar = (v_apagar / max_val) * 100
        label_pago = ""
        if v_pago > 0:
            pos = "bar-label-middle" if h_pago >= 50 else "bar-label-top"
            label_pago = f'<span class="bar-label {pos}">{fmt_money(v_pago)}</span>'
        label_apagar = ""
        if v_apagar > 0:
            pos = "bar-label-middle" if h_apagar >= 50 else "bar-label-top"
            label_apagar = f'<span class="bar-label {pos}">{fmt_money(v_apagar)}</span>'
        mes_label = f"{MESES[mes - 1]}/{str(ano)[-2:]}"
        barras_html.append(f'''
                <div class="bar-group">
                    <div class="bar-pair">
                        <div class="bar bar-pago" style="height: {h_pago}%;">
                            {label_pago}
                        </div>
                        <div class="bar bar-a-pagar" style="height: {h_apagar}%;">
                            {label_apagar}
                        </div>
                    </div>
                    <div class="bar-month">{mes_label}</div>
                </div>''')
    barras_html = "".join(barras_html)

    # --- Custo por etapa ---
    etapa_valor = defaultdict(float)
    for p in pagamentos:
        if p["etapa"]:
            etapa_valor[p["etapa"]] += p["valor"]
    etapas_ordenadas = [e for e in ordem_etapas if etapa_valor.get(e, 0) > 0]
    for e in etapa_valor:
        if e not in etapas_ordenadas:
            etapas_ordenadas.append(e)
    total_etapas = sum(etapa_valor.values()) or 1.0

    etapas_html = []
    for e in etapas_ordenadas:
        val = etapa_valor[e]
        pct = val / total_etapas * 100
        cor = etapas_cor.get(e, "#999999")
        ancora = slug_etapa(e)
        etapas_html.append(f'''
                    <a href="#{ancora}" class="etapa-item etapa-item-link" title="Ver lançamentos de {e}">
                        <div class="etapa-color" style="background: {cor};"></div>
                        <div class="etapa-info">
                            <div class="etapa-nome">{e} <span class="etapa-ver-mais">ver lançamentos →</span></div>
                            <div class="etapa-valores">
                                <span class="etapa-valor">{fmt_money(val)}</span>
                                <span class="etapa-pct">({pct:.1f}%)</span>
                            </div>
                        </div>
                        <div class="etapa-bar-bg">
                            <div class="etapa-bar-fill" style="width: {pct}%; background: {cor};"></div>
                        </div>
                    </a>''')
    etapas_html = "".join(etapas_html)

    # --- A pagar por fornecedor ---
    fornecedor_rows = defaultdict(list)
    for p in pagamentos:
        if p["status"] == "A Pagar":
            fornecedor_rows[p["fornecedor"] or "(sem fornecedor)"].append(p)

    fornecedores_ordenados = sorted(
        fornecedor_rows.items(), key=lambda kv: sum(x["valor"] for x in kv[1]), reverse=True
    )

    apagar_html = []
    for fornecedor, itens in fornecedores_ordenados:
        itens_ordenados = sorted(itens, key=lambda x: x["vencimento"] or datetime.date.max)
        total_forn = sum(x["valor"] for x in itens_ordenados)
        if len(itens_ordenados) == 1:
            item = itens_ordenados[0]
            venc = item["vencimento"].strftime("%d/%m/%Y") if item["vencimento"] else "-"
            apagar_html.append(f'''
                    <tr class="linha-unica">
                        <td class="fornecedor" data-label="Fornecedor">{fornecedor}</td>
                        <td class="vencimento" data-label="Vencimento">{venc}</td>
                        <td class="valor" data-label="Valor">{fmt_money(item["valor"])}</td>
                    </tr>''')
        else:
            apagar_html.append(f'''
                    <tr class="linha-fornecedor">
                        <td class="fornecedor" data-label="Fornecedor">{fornecedor}</td>
                        <td class="vencimento" data-label="Vencimento">{len(itens_ordenados)} parcelas</td>
                        <td class="valor" data-label="Valor">{fmt_money(total_forn)}</td>
                    </tr>''')
            for item in itens_ordenados:
                venc = item["vencimento"].strftime("%d/%m/%Y") if item["vencimento"] else "-"
                apagar_html.append(f'''
                    <tr class="linha-parcela">
                        <td class="fornecedor-nome" data-label="Fornecedor">{item["descricao"] or fornecedor}</td>
                        <td class="vencimento" data-label="Vencimento">{venc}</td>
                        <td class="valor" data-label="Valor">{fmt_money(item["valor"])}</td>
                    </tr>''')
    apagar_html = "".join(apagar_html) if apagar_html else (
        '<tr class="linha-unica"><td colspan="3" style="text-align:center;color:#999;">'
        'Nenhum lançamento em aberto</td></tr>'
    )

    # --- Pago por fornecedor ---
    pago_fornecedor_rows = defaultdict(list)
    for p in pagamentos:
        if p["status"] == "Pago":
            pago_fornecedor_rows[p["fornecedor"] or "(sem fornecedor)"].append(p)

    pago_fornecedores_ordenados = sorted(
        pago_fornecedor_rows.items(), key=lambda kv: sum(x["valor"] for x in kv[1]), reverse=True
    )

    pago_html = []
    for fornecedor, itens in pago_fornecedores_ordenados:
        itens_ordenados = sorted(itens, key=lambda x: x["data"] or datetime.date.min, reverse=True)
        total_forn = sum(x["valor"] for x in itens_ordenados)
        if len(itens_ordenados) == 1:
            item = itens_ordenados[0]
            data = item["data"].strftime("%d/%m/%Y") if item["data"] else "-"
            pago_html.append(f'''
                    <tr class="linha-unica-pago">
                        <td class="fornecedor" data-label="Fornecedor">{fornecedor}</td>
                        <td class="vencimento" data-label="Data">{data}</td>
                        <td class="valor" data-label="Valor">{fmt_money(item["valor"])}</td>
                    </tr>''')
        else:
            pago_html.append(f'''
                    <tr class="linha-fornecedor-pago">
                        <td class="fornecedor" data-label="Fornecedor">{fornecedor}</td>
                        <td class="vencimento" data-label="Data">{len(itens_ordenados)} pagamentos</td>
                        <td class="valor" data-label="Valor">{fmt_money(total_forn)}</td>
                    </tr>''')
            for item in itens_ordenados:
                data = item["data"].strftime("%d/%m/%Y") if item["data"] else "-"
                pago_html.append(f'''
                    <tr class="linha-parcela-pago">
                        <td class="fornecedor-nome" data-label="Fornecedor">{item["descricao"] or fornecedor}</td>
                        <td class="vencimento" data-label="Data">{data}</td>
                        <td class="valor" data-label="Valor">{fmt_money(item["valor"])}</td>
                    </tr>''')
    pago_html = "".join(pago_html) if pago_html else (
        '<tr class="linha-unica-pago"><td colspan="3" style="text-align:center;color:#999;">'
        'Nenhum pagamento registrado</td></tr>'
    )

    # --- Materiais: quantidade total consolidada por produto (todas as etapas) ---
    mat_total = defaultdict(lambda: {"quantidade": 0.0, "custo_total": 0.0, "etapas": set()})
    for m in materiais:
        agg = mat_total[m["produto"]]
        agg["quantidade"] += m["quantidade"]
        agg["custo_total"] += m["custo_total"]
        if m["etapa"]:
            agg["etapas"].add(m["etapa"])
    mat_total_ordenado = sorted(mat_total.items(), key=lambda kv: kv[0].upper())

    mat_total_html = []
    for produto, agg in mat_total_ordenado:
        custo_und = agg["custo_total"] / agg["quantidade"] if agg["quantidade"] else 0.0
        n_etapas = len(agg["etapas"])
        etapas_txt = ", ".join(sorted(agg["etapas"])) if n_etapas <= 2 else f"{n_etapas} etapas"
        produto_attr = produto.lower().replace('"', "&quot;")
        classe_alerta = " linha-verificar" if produto == "Verificar" else ""
        icone = "⚠️ " if produto == "Verificar" else ""
        mat_total_html.append(f'''
                    <tr class="linha-material{classe_alerta}" data-produto="{produto_attr}">
                        <td class="mat-produto" data-label="Produto">{icone}{produto}</td>
                        <td class="mat-num mat-qtd-total" data-label="Qtd. Total">{agg["quantidade"]:.2f}</td>
                        <td class="mat-num" data-label="Custo UND Médio">{fmt_money(custo_und)}</td>
                        <td class="mat-num mat-total" data-label="Custo Total">{fmt_money(agg["custo_total"])}</td>
                        <td class="mat-etapa" data-label="Etapa(s)">{etapas_txt or "-"}</td>
                    </tr>''')
    mat_total_html = "".join(mat_total_html) if mat_total_html else (
        '<tr><td colspan="5" style="text-align:center;color:#999;">Nenhum material lançado</td></tr>'
    )

    # --- Materiais agregados (por produto + etapa) ---
    mat_agrupado = defaultdict(lambda: {"quantidade": 0.0, "custo_total": 0.0, "etapa": ""})
    for m in materiais:
        chave = (m["produto"], m["etapa"])
        agg = mat_agrupado[chave]
        agg["quantidade"] += m["quantidade"]
        agg["custo_total"] += m["custo_total"]
        agg["etapa"] = m["etapa"]
    materiais_ordenados = sorted(mat_agrupado.items(), key=lambda kv: kv[1]["custo_total"], reverse=True)

    materiais_html = []
    for (produto, _etapa_chave), agg in materiais_ordenados:
        custo_und = agg["custo_total"] / agg["quantidade"] if agg["quantidade"] else 0.0
        etapa_txt = agg["etapa"] or "-"
        etapa_cor = etapas_cor.get(agg["etapa"], "#cbd5e0")
        produto_attr = produto.lower().replace('"', "&quot;")
        classe_alerta = " linha-verificar" if produto == "Verificar" else ""
        icone = "⚠️ " if produto == "Verificar" else ""
        materiais_html.append(f'''
                    <tr class="linha-material{classe_alerta}" data-produto="{produto_attr}">
                        <td class="mat-produto" data-label="Produto">{icone}{produto}</td>
                        <td class="mat-etapa" data-label="Etapa"><span class="mat-etapa-dot" style="background:{etapa_cor};"></span>{etapa_txt}</td>
                        <td class="mat-num" data-label="Quantidade">{agg["quantidade"]:.2f}</td>
                        <td class="mat-num" data-label="Custo UND">{fmt_money(custo_und)}</td>
                        <td class="mat-num mat-total" data-label="Custo Total">{fmt_money(agg["custo_total"])}</td>
                    </tr>''')
    materiais_html = "".join(materiais_html) if materiais_html else (
        '<tr><td colspan="5" style="text-align:center;color:#999;">Nenhum material lançado</td></tr>'
    )

    # --- Por Responsável ---
    PALETA_RESP = ["#4472C4", "#ED7D31", "#70AD47", "#9E480E", "#5B9BD5", "#FFC000", "#7F8C8D", "#8E44AD"]
    resp_agg = defaultdict(lambda: {"pago": 0.0, "a_pagar": 0.0, "itens": []})
    for p in pagamentos:
        nome_resp = p["responsavel"] or "Não informado"
        agg = resp_agg[nome_resp]
        if p["status"] == "Pago":
            agg["pago"] += p["valor"]
        elif p["status"] == "A Pagar":
            agg["a_pagar"] += p["valor"]
        agg["itens"].append(p)

    resp_ordenados = sorted(
        resp_agg.items(), key=lambda kv: kv[1]["pago"] + kv[1]["a_pagar"], reverse=True
    )
    max_resp_total = max([v["pago"] + v["a_pagar"] for _, v in resp_ordenados], default=0.0) or 1.0

    resp_cores = {}
    idx_cor = 0
    for nome, _ in resp_ordenados:
        if nome != "Não informado":
            resp_cores[nome] = PALETA_RESP[idx_cor % len(PALETA_RESP)]
            idx_cor += 1
    resp_cores["Não informado"] = "#bdc3c7"

    resp_resumo_html = []
    resp_detalhe_html = []
    for nome, agg in resp_ordenados:
        total_nome = agg["pago"] + agg["a_pagar"]
        pct = total_nome / max_resp_total * 100
        cor = resp_cores[nome]
        resp_resumo_html.append(f'''
                    <div class="etapa-item">
                        <div class="etapa-color" style="background: {cor};"></div>
                        <div class="etapa-info">
                            <div class="etapa-nome">{nome}</div>
                            <div class="etapa-valores">
                                <span class="etapa-valor">{fmt_money(total_nome)}</span>
                                <span class="etapa-pct">(Pago: {fmt_money(agg["pago"])} · A Pagar: {fmt_money(agg["a_pagar"])})</span>
                            </div>
                        </div>
                        <div class="etapa-bar-bg">
                            <div class="etapa-bar-fill" style="width: {pct}%; background: {cor};"></div>
                        </div>
                    </div>''')

        linhas_itens = []
        itens_ordenados = sorted(
            agg["itens"], key=lambda x: x["data"] or x["vencimento"] or datetime.date.min, reverse=True
        )
        for item in itens_ordenados:
            data_ref = item["data"] or item["vencimento"]
            data_txt = data_ref.strftime("%d/%m/%Y") if data_ref else "-"
            status_classe = "resp-status-pago" if item["status"] == "Pago" else "resp-status-apagar"
            linhas_itens.append(f'''
                        <tr>
                            <td data-label="Fornecedor">{item["fornecedor"] or "-"}</td>
                            <td data-label="Etapa">{item["etapa"] or "-"}</td>
                            <td data-label="Data">{data_txt}</td>
                            <td data-label="Status"><span class="resp-status {status_classe}">{item["status"] or "-"}</span></td>
                            <td class="mat-num" data-label="Valor">{fmt_money(item["valor"])}</td>
                        </tr>''')
        resp_detalhe_html.append(f'''
                <div class="resp-bloco">
                    <div class="resp-bloco-titulo">
                        <span class="etapa-color" style="background:{cor};"></span>
                        {nome}
                        <span class="resp-bloco-total">{fmt_money(total_nome)}</span>
                    </div>
                    <div class="table-wrapper" style="border:none; border-radius:0;">
                    <table class="materiais-table tabela-empilhavel">
                        <thead>
                            <tr><th>Fornecedor</th><th>Etapa</th><th>Data</th><th>Status</th><th>Valor</th></tr>
                        </thead>
                        <tbody>{"".join(linhas_itens)}
                        </tbody>
                    </table>
                    </div>
                </div>''')

    resp_resumo_html = "".join(resp_resumo_html)
    resp_detalhe_html = "".join(resp_detalhe_html)

    # --- Compras por Etapa ---
    etapa_agg = defaultdict(lambda: {"pago": 0.0, "a_pagar": 0.0, "itens": []})
    for p in pagamentos:
        nome_etapa = p["etapa"] or "Sem etapa"
        agg = etapa_agg[nome_etapa]
        if p["status"] == "Pago":
            agg["pago"] += p["valor"]
        elif p["status"] == "A Pagar":
            agg["a_pagar"] += p["valor"]
        agg["itens"].append(p)

    etapa_agg_ordenado = sorted(
        etapa_agg.items(), key=lambda kv: kv[1]["pago"] + kv[1]["a_pagar"], reverse=True
    )
    max_etapa_compra = max([v["pago"] + v["a_pagar"] for _, v in etapa_agg_ordenado], default=0.0) or 1.0

    compras_resumo_html = []
    compras_detalhe_html = []
    for nome_etapa, agg in etapa_agg_ordenado:
        total_etapa_compra = agg["pago"] + agg["a_pagar"]
        pct = total_etapa_compra / max_etapa_compra * 100
        cor = etapas_cor.get(nome_etapa, "#999999") if nome_etapa != "Sem etapa" else "#bdc3c7"
        ancora_etapa = slug_etapa(nome_etapa)
        compras_resumo_html.append(f'''
                    <a href="#{ancora_etapa}" class="etapa-item etapa-item-link" title="Ver lançamentos de {nome_etapa}">
                        <div class="etapa-color" style="background: {cor};"></div>
                        <div class="etapa-info">
                            <div class="etapa-nome">{nome_etapa} <span class="etapa-ver-mais">ver lançamentos →</span></div>
                            <div class="etapa-valores">
                                <span class="etapa-valor">{fmt_money(total_etapa_compra)}</span>
                                <span class="etapa-pct">(Pago: {fmt_money(agg["pago"])} · A Pagar: {fmt_money(agg["a_pagar"])})</span>
                            </div>
                        </div>
                        <div class="etapa-bar-bg">
                            <div class="etapa-bar-fill" style="width: {pct}%; background: {cor};"></div>
                        </div>
                    </a>''')

        itens_ordenados = sorted(
            agg["itens"], key=lambda x: x["data"] or x["vencimento"] or datetime.date.min, reverse=True
        )
        linhas_compra = []
        for item in itens_ordenados:
            data_ref = item["data"] or item["vencimento"]
            data_txt = data_ref.strftime("%d/%m/%Y") if data_ref else "-"
            status_classe = "resp-status-pago" if item["status"] == "Pago" else "resp-status-apagar"
            linhas_compra.append(f'''
                        <tr>
                            <td data-label="Fornecedor">{item["fornecedor"] or "-"}</td>
                            <td data-label="Produto">{item["descricao"] or "-"}</td>
                            <td data-label="Data">{data_txt}</td>
                            <td data-label="Status"><span class="resp-status {status_classe}">{item["status"] or "-"}</span></td>
                            <td class="mat-num" data-label="Valor">{fmt_money(item["valor"])}</td>
                        </tr>''')
        compras_detalhe_html.append(f'''
                <div class="resp-bloco compra-etapa-bloco" id="{slug_etapa(nome_etapa)}">
                    <div class="resp-bloco-titulo">
                        <span class="etapa-color" style="background:{cor};"></span>
                        {nome_etapa}
                        <span class="resp-bloco-total">{fmt_money(total_etapa_compra)}</span>
                    </div>
                    <div class="table-wrapper" style="border:none; border-radius:0;">
                    <table class="materiais-table tabela-empilhavel">
                        <thead>
                            <tr><th>Fornecedor</th><th>Produto</th><th>Data</th><th>Status</th><th>Valor</th></tr>
                        </thead>
                        <tbody>{"".join(linhas_compra)}
                        </tbody>
                    </table>
                    </div>
                </div>''')

    compras_resumo_html = "".join(compras_resumo_html)
    compras_detalhe_html = "".join(compras_detalhe_html)

    # --- Por Tipo de Gasto (Material, Mão de Obra, Equipamento...) ---
    PALETA_TIPO = ["#16A085", "#795548", "#C0392B", "#8E44AD", "#2980B9", "#D35400", "#7F8C8D"]
    tipo_agg = defaultdict(lambda: {"pago": 0.0, "a_pagar": 0.0, "itens": []})
    for p in pagamentos:
        nome_tipo = p["tipo_gasto"] or "Não classificado"
        agg = tipo_agg[nome_tipo]
        if p["status"] == "Pago":
            agg["pago"] += p["valor"]
        elif p["status"] == "A Pagar":
            agg["a_pagar"] += p["valor"]
        agg["itens"].append(p)

    tipo_agg_ordenado = sorted(
        tipo_agg.items(), key=lambda kv: kv[1]["pago"] + kv[1]["a_pagar"], reverse=True
    )
    max_tipo_total = max([v["pago"] + v["a_pagar"] for _, v in tipo_agg_ordenado], default=0.0) or 1.0

    tipo_cores = {}
    idx_cor_tipo = 0
    for nome_tipo, _ in tipo_agg_ordenado:
        if nome_tipo != "Não classificado":
            tipo_cores[nome_tipo] = PALETA_TIPO[idx_cor_tipo % len(PALETA_TIPO)]
            idx_cor_tipo += 1
    tipo_cores["Não classificado"] = "#bdc3c7"

    tipo_resumo_html = []
    tipo_detalhe_html = []
    for nome_tipo, agg in tipo_agg_ordenado:
        total_tipo = agg["pago"] + agg["a_pagar"]
        pct = total_tipo / max_tipo_total * 100
        cor = tipo_cores[nome_tipo]
        ancora_tipo = "tipo-" + slug_etapa(nome_tipo).replace("etapa-", "")
        tipo_resumo_html.append(f'''
                    <a href="#{ancora_tipo}" class="etapa-item etapa-item-link" title="Ver lançamentos de {nome_tipo}">
                        <div class="etapa-color" style="background: {cor};"></div>
                        <div class="etapa-info">
                            <div class="etapa-nome">{nome_tipo} <span class="etapa-ver-mais">ver lançamentos →</span></div>
                            <div class="etapa-valores">
                                <span class="etapa-valor">{fmt_money(total_tipo)}</span>
                                <span class="etapa-pct">(Pago: {fmt_money(agg["pago"])} · A Pagar: {fmt_money(agg["a_pagar"])})</span>
                            </div>
                        </div>
                        <div class="etapa-bar-bg">
                            <div class="etapa-bar-fill" style="width: {pct}%; background: {cor};"></div>
                        </div>
                    </a>''')

        itens_ordenados = sorted(
            agg["itens"], key=lambda x: x["data"] or x["vencimento"] or datetime.date.min, reverse=True
        )
        linhas_tipo = []
        for item in itens_ordenados:
            data_ref = item["data"] or item["vencimento"]
            data_txt = data_ref.strftime("%d/%m/%Y") if data_ref else "-"
            status_classe = "resp-status-pago" if item["status"] == "Pago" else "resp-status-apagar"
            linhas_tipo.append(f'''
                        <tr>
                            <td data-label="Fornecedor">{item["fornecedor"] or "-"}</td>
                            <td data-label="Produto">{item["descricao"] or "-"}</td>
                            <td data-label="Etapa">{item["etapa"] or "-"}</td>
                            <td data-label="Data">{data_txt}</td>
                            <td data-label="Status"><span class="resp-status {status_classe}">{item["status"] or "-"}</span></td>
                            <td class="mat-num" data-label="Valor">{fmt_money(item["valor"])}</td>
                        </tr>''')
        tipo_detalhe_html.append(f'''
                <div class="resp-bloco tipo-gasto-bloco" id="{ancora_tipo}">
                    <div class="resp-bloco-titulo">
                        <span class="etapa-color" style="background:{cor};"></span>
                        {nome_tipo}
                        <span class="resp-bloco-total">{fmt_money(total_tipo)}</span>
                    </div>
                    <div class="table-wrapper" style="border:none; border-radius:0;">
                    <table class="materiais-table tabela-empilhavel">
                        <thead>
                            <tr><th>Fornecedor</th><th>Produto</th><th>Etapa</th><th>Data</th><th>Status</th><th>Valor</th></tr>
                        </thead>
                        <tbody>{"".join(linhas_tipo)}
                        </tbody>
                    </table>
                    </div>
                </div>''')

    tipo_resumo_html = "".join(tipo_resumo_html)
    tipo_detalhe_html = "".join(tipo_detalhe_html)

    subtitulo = f"{MESES[hoje.month - 1]}/{hoje.year}"

    html = f'''<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Relatório {obra_id} - Gestão de Obras</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); min-height: 100vh; padding: 10px; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; border-radius: 20px; overflow: hidden; box-shadow: 0 20px 60px rgba(0,0,0,0.3); }}
        .header {{ background: linear-gradient(135deg, #2c3e50 0%, #34495e 100%); color: white; padding: 20px; text-align: center; }}
        .header h1 {{ font-size: 22px; margin-bottom: 5px; font-weight: 600; }}
        .header .subtitle {{ font-size: 12px; opacity: 0.8; }}
        .content {{ padding: 20px; }}
        .tabs-nav {{ display: flex; gap: 8px; padding: 0 20px; background: #2c3e50; }}
        .tab-btn {{ flex: 1; display: block; background: transparent; border: none; color: rgba(255,255,255,0.65); padding: 14px 10px; font-size: 14px; font-weight: 600; cursor: pointer; border-bottom: 3px solid transparent; transition: all 0.15s; font-family: inherit; text-align: center; text-decoration: none; user-select: none; }}
        .tab-btn:hover {{ color: white; }}
        /* Sistema de abas 100% em CSS (âncoras + :target), sem JavaScript.
           Por padrão, sem nenhuma âncora na URL, a aba "Visão Geral" fica visível. */
        .tab-panel {{ display: none; }}
        #tab-geral {{ display: block; }}
        .tabs-nav a[href="#tab-geral"] {{ color: white; border-bottom-color: #ED7D31; }}

        body:has(#tab-pagamentos:target) #tab-geral {{ display: none; }}
        body:has(#tab-pagamentos:target) #tab-pagamentos {{ display: block; }}
        body:has(#tab-pagamentos:target) .tabs-nav a[href="#tab-geral"] {{ color: rgba(255,255,255,0.65); border-bottom-color: transparent; }}
        body:has(#tab-pagamentos:target) .tabs-nav a[href="#tab-pagamentos"] {{ color: white; border-bottom-color: #ED7D31; }}

        body:has(#tab-responsavel:target) #tab-geral {{ display: none; }}
        body:has(#tab-responsavel:target) #tab-responsavel {{ display: block; }}
        body:has(#tab-responsavel:target) .tabs-nav a[href="#tab-geral"] {{ color: rgba(255,255,255,0.65); border-bottom-color: transparent; }}
        body:has(#tab-responsavel:target) .tabs-nav a[href="#tab-responsavel"] {{ color: white; border-bottom-color: #ED7D31; }}

        /* Aba "Compras por Etapa": abre tanto pelo botão da aba quanto ao clicar
           numa etapa na Visão Geral (que pula direto pro bloco daquela etapa). */
        body:has(#tab-compras-etapa:target) #tab-geral,
        body:has(.compra-etapa-bloco:target) #tab-geral {{
            display: none;
        }}
        body:has(#tab-compras-etapa:target) #tab-compras-etapa,
        body:has(.compra-etapa-bloco:target) #tab-compras-etapa {{
            display: block;
        }}
        body:has(#tab-compras-etapa:target) .tabs-nav a[href="#tab-geral"],
        body:has(.compra-etapa-bloco:target) .tabs-nav a[href="#tab-geral"] {{
            color: rgba(255,255,255,0.65); border-bottom-color: transparent;
        }}
        body:has(#tab-compras-etapa:target) .tabs-nav a[href="#tab-compras-etapa"],
        body:has(.compra-etapa-bloco:target) .tabs-nav a[href="#tab-compras-etapa"] {{
            color: white; border-bottom-color: #ED7D31;
        }}
        .compra-etapa-bloco {{ scroll-margin-top: 16px; }}
        .compra-etapa-bloco:target {{
            outline: 3px solid #ED7D31; outline-offset: 2px; animation: piscaEtapa 1.6s ease-out 1;
        }}
        /* Aba "Tipo de Gasto": mesmo esquema, abre pelo botão ou ao clicar num tipo. */
        body:has(#tab-tipo-gasto:target) #tab-geral,
        body:has(.tipo-gasto-bloco:target) #tab-geral {{
            display: none;
        }}
        body:has(#tab-tipo-gasto:target) #tab-tipo-gasto,
        body:has(.tipo-gasto-bloco:target) #tab-tipo-gasto {{
            display: block;
        }}
        body:has(#tab-tipo-gasto:target) .tabs-nav a[href="#tab-geral"],
        body:has(.tipo-gasto-bloco:target) .tabs-nav a[href="#tab-geral"] {{
            color: rgba(255,255,255,0.65); border-bottom-color: transparent;
        }}
        body:has(#tab-tipo-gasto:target) .tabs-nav a[href="#tab-tipo-gasto"],
        body:has(.tipo-gasto-bloco:target) .tabs-nav a[href="#tab-tipo-gasto"] {{
            color: white; border-bottom-color: #ED7D31;
        }}
        .tipo-gasto-bloco {{ scroll-margin-top: 16px; }}
        .tipo-gasto-bloco:target {{
            outline: 3px solid #ED7D31; outline-offset: 2px; animation: piscaEtapa 1.6s ease-out 1;
        }}
        @keyframes piscaEtapa {{
            0% {{ background: #fff3cd; }}
            100% {{ background: transparent; }}
        }}
        .etapa-item-link {{ text-decoration: none; color: inherit; cursor: pointer; transition: transform 0.1s; }}
        .etapa-item-link:hover {{ transform: translateX(2px); }}
        .etapa-item-link:hover .etapa-nome {{ text-decoration: underline; }}
        .etapa-ver-mais {{ font-size: 11px; font-weight: 400; color: #8492a6; margin-left: 6px; white-space: nowrap; }}
        .resp-bloco {{ margin-bottom: 22px; background: white; border-radius: 10px; overflow: hidden; border: 1px solid #e2e8f0; }}
        .resp-bloco-titulo {{ display: flex; align-items: center; gap: 10px; background: #f7fafc; padding: 12px 15px; font-size: 15px; font-weight: 700; color: #2d3748; }}
        .resp-bloco-titulo .etapa-color {{ width: 14px; height: 14px; margin-right: 0; }}
        .resp-bloco-total {{ margin-left: auto; color: #2980b9; font-size: 16px; }}
        .resp-status {{ font-size: 11px; font-weight: 700; padding: 2px 8px; border-radius: 10px; text-transform: uppercase; }}
        .resp-status-pago {{ background: #f0fff4; color: #2f855a; }}
        .resp-status-apagar {{ background: #fff3cd; color: #b7791f; }}
        .section-title {{ background: #95a5a6; color: white; padding: 12px; text-align: center; font-size: 16px; font-weight: 600; margin: 20px 0 15px 0; border-radius: 8px; }}
        .cards-container {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); gap: 12px; margin-bottom: 15px; }}
        .card {{ background: #f8f9fa; border-radius: 12px; padding: 18px 12px; text-align: center; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
        .card-icon {{ font-size: 32px; margin-bottom: 8px; }}
        .card-title {{ font-size: 11px; color: #666; margin-bottom: 8px; font-weight: 600; text-transform: uppercase; }}
        .card-value {{ font-size: 20px; font-weight: 700; color: #2980b9; }}
        .alert-card {{ background: #fff3cd; border: 2px solid #ffc107; grid-column: 1 / -1; }}
        .alert-card .card-icon {{ color: #ff9800; }}
        .alert-card .card-value {{ color: #d32f2f; font-size: 22px; }}
        .chart-container {{ background: white; border-radius: 12px; padding: 20px 15px; margin: 15px 0; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
        .chart-title {{ font-size: 16px; font-weight: 600; color: #2c3e50; text-align: center; margin-bottom: 15px; }}
        .chart-legend {{ display: flex; justify-content: center; gap: 20px; margin-bottom: 15px; flex-wrap: wrap; }}
        .legend-item {{ display: flex; align-items: center; gap: 6px; font-size: 12px; }}
        .legend-color {{ width: 16px; height: 16px; border-radius: 3px; }}
        .legend-pago {{ background: #4472C4; }}
        .legend-a-pagar {{ background: #ED7D31; }}
        .bars-container {{ display: flex; justify-content: space-around; align-items: flex-end; height: 250px; gap: 8px; padding: 10px; }}
        .bar-group {{ flex: 1; display: flex; flex-direction: column; align-items: center; min-width: 60px; }}
        .bar-pair {{ display: flex; gap: 4px; height: 200px; width: 100%; justify-content: center; align-items: flex-end; }}
        .bar {{ width: 28px; border-radius: 4px 4px 0 0; position: relative; min-height: 3px; }}
        .bar-pago {{ background: #4472C4; }}
        .bar-a-pagar {{ background: #ED7D31; }}
        .bar-label-top {{ position: absolute; bottom: 100%; left: 50%; transform: translateX(-50%); font-size: 9px; font-weight: 600; color: #2c3e50; white-space: nowrap; margin-bottom: 4px; }}
        .bar-label-middle {{ position: absolute; top: 50%; left: 50%; transform: translate(-50%, -50%); font-size: 10px; font-weight: 700; color: #2c3e50; background: rgba(255,255,255,0.9); padding: 2px 6px; border-radius: 4px; white-space: nowrap; }}
        .bar-month {{ margin-top: 8px; font-size: 11px; font-weight: 600; color: #34495e; }}
        .etapas-container {{ background: #f8f9fa; border-radius: 12px; padding: 15px; }}
        .etapa-item {{ margin-bottom: 15px; display: block; }}
        .etapa-color {{ width: 20px; height: 20px; border-radius: 4px; display: inline-block; margin-right: 10px; }}
        .etapa-info {{ display: flex; justify-content: space-between; margin-bottom: 6px; }}
        .etapa-nome {{ font-size: 14px; font-weight: 600; color: #2c3e50; }}
        .etapa-valores {{ font-size: 13px; color: #7f8c8d; }}
        .etapa-valor {{ font-weight: 600; color: #2c3e50; margin-right: 8px; }}
        .etapa-pct {{ color: #95a5a6; }}
        .etapa-bar-bg {{ background: #ecf0f1; height: 24px; border-radius: 12px; overflow: hidden; }}
        .etapa-bar-fill {{ height: 100%; border-radius: 12px; }}
        .totais-container {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin: 20px 0; }}
        .total-card {{ text-align: center; padding: 20px; border-radius: 8px; border: 2px solid; }}
        .total-card.acumulado {{ background: #e8f4f8; border-color: #3498db; }}
        .total-card.quitado {{ background: #f0fff4; border-color: #38a169; }}
        .total-label {{ font-size: 13px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 8px; }}
        .total-card.acumulado .total-label {{ color: #2c5aa0; }}
        .total-card.quitado .total-label {{ color: #2f855a; }}
        .total-value {{ font-size: 28px; font-weight: 700; }}
        .total-card.acumulado .total-value {{ color: #2c3e50; }}
        .total-card.quitado .total-value {{ color: #38a169; }}
        .a-pagar-section {{ background: white; border-radius: 12px; padding: 20px; margin: 20px 0; box-shadow: 0 2px 8px rgba(0,0,0,0.08); border-left: 4px solid #ed8936; }}
        .section-title-a-pagar {{ font-size: 16px; font-weight: 700; color: #2d3748; margin-bottom: 16px; display: flex; align-items: center; justify-content: space-between; }}
        .total-a-pagar {{ font-size: 20px; font-weight: 700; color: #ed8936; }}
        .tabela-a-pagar {{ width: 100%; border-collapse: collapse; }}
        .tabela-a-pagar thead {{ background: #f7fafc; border-bottom: 2px solid #e1e8ed; }}
        .tabela-a-pagar th {{ padding: 12px; text-align: left; font-size: 12px; font-weight: 600; color: #718096; text-transform: uppercase; letter-spacing: 0.5px; }}
        .tabela-a-pagar th:last-child {{ text-align: right; }}
        .tabela-a-pagar td {{ padding: 12px; font-size: 14px; border-bottom: 1px solid #f1f3f5; }}
        .tabela-a-pagar .valor {{ text-align: right; }}
        .linha-unica td {{ background: white; font-weight: 600; color: #2d3748; }}
        .linha-unica .fornecedor {{ color: #2d3748; }}
        .linha-unica .vencimento {{ color: #4a5568; }}
        .linha-unica .valor {{ color: #ed8936; }}
        .linha-fornecedor td {{ background: #fef5e7; font-weight: 700; color: #2d3748; border-bottom: 1px solid #fbd38d; }}
        .linha-fornecedor .vencimento {{ color: #718096; font-weight: 400; font-size: 12px; font-style: italic; }}
        .linha-fornecedor .valor {{ color: #c05621; }}
        .linha-parcela td {{ background: white; color: #4a5568; font-size: 13px; border-bottom: 1px solid #f7fafc; }}
        .linha-parcela .fornecedor-nome {{ padding-left: 30px; color: #718096; }}
        .linha-parcela .fornecedor-nome::before {{ content: "↳ "; color: #cbd5e0; }}
        .linha-parcela .vencimento {{ color: #4a5568; }}
        .linha-parcela .valor {{ color: #ed8936; font-weight: 500; }}
        .pago-section {{ background: white; border-radius: 12px; padding: 20px; margin: 20px 0; box-shadow: 0 2px 8px rgba(0,0,0,0.08); border-left: 4px solid #38a169; }}
        .total-pago-fornecedor {{ font-size: 20px; font-weight: 700; color: #38a169; }}
        .linha-unica-pago td {{ background: white; font-weight: 600; color: #2d3748; }}
        .linha-unica-pago .fornecedor {{ color: #2d3748; }}
        .linha-unica-pago .vencimento {{ color: #4a5568; }}
        .linha-unica-pago .valor {{ color: #38a169; }}
        .linha-fornecedor-pago td {{ background: #f0fff4; font-weight: 700; color: #2d3748; border-bottom: 1px solid #9ae6b4; }}
        .linha-fornecedor-pago .vencimento {{ color: #718096; font-weight: 400; font-size: 12px; font-style: italic; }}
        .linha-fornecedor-pago .valor {{ color: #2f855a; }}
        .linha-parcela-pago td {{ background: white; color: #4a5568; font-size: 13px; border-bottom: 1px solid #f7fafc; }}
        .linha-parcela-pago .fornecedor-nome {{ padding-left: 30px; color: #718096; }}
        .linha-parcela-pago .fornecedor-nome::before {{ content: "↳ "; color: #cbd5e0; }}
        .linha-parcela-pago .vencimento {{ color: #4a5568; }}
        .linha-parcela-pago .valor {{ color: #38a169; font-weight: 500; }}
        .materiais-section {{ margin-top: 20px; }}
        .materiais-title {{ background: #2c3e50; color: white; padding: 10px 15px; border-radius: 8px 8px 0 0; font-size: 14px; font-weight: 600; }}
        .table-wrapper {{ overflow-x: auto; border-radius: 0 0 8px 8px; border: 1px solid #dce1e7; }}
        .materiais-table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
        .materiais-table thead tr {{ background: #34495e; color: white; }}
        .materiais-table th {{ padding: 10px 12px; text-align: left; font-weight: 600; white-space: nowrap; }}
        .materiais-table th:not(:first-child) {{ text-align: right; }}
        .materiais-table tbody tr:nth-child(even) {{ background: #f4f6f8; }}
        .materiais-table tbody tr:hover {{ background: #eaf3fb; }}
        .materiais-table td {{ padding: 9px 12px; border-bottom: 1px solid #e9ecef; color: #2c3e50; }}
        .mat-produto {{ max-width: 200px; word-break: break-word; }}
        .mat-etapa {{ white-space: nowrap; color: #4a5568; font-size: 12px; }}
        .mat-etapa-dot {{ display: inline-block; width: 8px; height: 8px; border-radius: 50%; margin-right: 6px; }}
        .mat-qtd-total {{ font-weight: 700; color: #2c3e50; }}
        .materiais-subtitle {{ font-size: 12px; color: #718096; margin: 4px 0 10px 0; text-align: center; }}
        .materiais-busca-wrap {{ display: flex; align-items: center; gap: 10px; margin: 20px 0 4px 0; }}
        .materiais-busca {{ flex: 1; padding: 10px 14px; border: 2px solid #dce1e7; border-radius: 8px; font-size: 16px; font-family: inherit; color: #2c3e50; -webkit-appearance: none; appearance: none; }}
        .materiais-busca:focus {{ outline: none; border-color: #4472C4; }}
        .materiais-busca-info {{ font-size: 12px; color: #718096; white-space: nowrap; }}
        .materiais-vazio {{ text-align: center; color: #999; padding: 14px; font-size: 13px; }}
        .aviso-sem-js {{ background: #fff3cd; border: 1px solid #ffc107; color: #7a5c00; border-radius: 8px; padding: 12px 14px; font-size: 13px; line-height: 1.5; margin: 8px 0 16px 0; }}
        .linha-verificar td {{ background: #fff8e1 !important; color: #8a6100; font-weight: 600; }}
        .linha-verificar:hover td {{ background: #ffedb3 !important; }}
        .mat-num {{ text-align: right; white-space: nowrap; }}
        .mat-total {{ font-weight: 700; color: #2980b9; }}
        .footer {{ text-align: center; padding: 15px; color: #7f8c8d; font-size: 11px; }}
        @media (max-width: 768px) {{ .totais-container {{ grid-template-columns: 1fr; }} }}

        /* ===== Ajustes para celular (retrato) ===== */
        @media (max-width: 600px) {{
            body {{ padding: 4px; }}
            .container {{ border-radius: 12px; }}
            .header {{ padding: 14px 10px; }}
            .header h1 {{ font-size: 17px; }}
            .header .subtitle {{ font-size: 11px; }}
            .content {{ padding: 10px; }}
            .tabs-nav {{ padding: 0 6px; }}
            .tab-btn {{ font-size: 10.5px; padding: 10px 2px; line-height: 1.25; }}
            .section-title {{ font-size: 13px; padding: 10px; margin: 14px 0 10px 0; }}
            .cards-container {{ grid-template-columns: 1fr 1fr; gap: 8px; }}
            .card {{ padding: 14px 8px; }}
            .card-icon {{ font-size: 24px; margin-bottom: 4px; }}
            .card-title {{ font-size: 10px; }}
            .card-value {{ font-size: 16px; }}
            .alert-card .card-value {{ font-size: 18px; }}
            .chart-container {{ padding: 14px 8px; }}
            .chart-title {{ font-size: 14px; }}
            .chart-legend {{ gap: 14px; margin-bottom: 10px; }}
            .bars-container {{
                height: auto; overflow-x: auto; overflow-y: hidden;
                justify-content: flex-start; padding: 6px 4px 10px 4px;
                -webkit-overflow-scrolling: touch;
            }}
            .bar-group {{ min-width: 46px; flex: 0 0 auto; }}
            .bar-pair {{ height: 150px; gap: 3px; }}
            .bar {{ width: 20px; }}
            .bar-label-top {{ font-size: 7px; margin-bottom: 2px; }}
            .bar-label-middle {{ font-size: 8px; padding: 1px 4px; }}
            .bar-month {{ font-size: 9px; margin-top: 6px; }}
            .etapas-container {{ padding: 10px; }}
            .etapa-item {{ margin-bottom: 12px; }}
            .etapa-nome {{ font-size: 13px; }}
            .etapa-info {{ flex-direction: column; gap: 2px; margin-bottom: 4px; }}
            .etapa-valores {{ font-size: 12px; }}
            .totais-container {{ gap: 10px; margin: 14px 0; }}
            .total-card {{ padding: 14px 8px; }}
            .total-label {{ font-size: 11px; }}
            .total-value {{ font-size: 20px; }}
            .a-pagar-section, .pago-section {{ padding: 12px; margin: 14px 0; }}
            .section-title-a-pagar {{ font-size: 13px; flex-wrap: wrap; gap: 4px; }}
            .total-a-pagar, .total-pago-fornecedor {{ font-size: 15px; }}
            .tabela-a-pagar th {{ font-size: 10px; padding: 8px 6px; }}
            .tabela-a-pagar td {{ font-size: 12px; padding: 8px 6px; }}
            .linha-parcela .fornecedor-nome {{ padding-left: 16px; }}
            .materiais-busca-wrap {{ flex-direction: column; align-items: stretch; gap: 6px; margin: 14px 0 4px 0; }}
            .materiais-busca {{ font-size: 16px; padding: 9px 12px; }}
            .materiais-busca-info {{ text-align: right; }}
            .materiais-title {{ font-size: 12px; padding: 8px 10px; }}
            .materiais-subtitle {{ font-size: 11px; }}
            .materiais-table th {{ font-size: 10px; padding: 7px 8px; }}
            .materiais-table td {{ font-size: 11px; padding: 7px 8px; }}
            .mat-produto {{ max-width: 110px; }}
            .mat-etapa {{ font-size: 10px; }}
            .resp-bloco-titulo {{ font-size: 12px; padding: 10px 12px; gap: 8px; }}
            .resp-bloco-total {{ font-size: 13px; }}
            .footer {{ font-size: 10px; }}
        }}
        @media (max-width: 360px) {{
            .cards-container {{ grid-template-columns: 1fr; }}
        }}
        /* Tabelas largas viram "cards" empilhados no celular — sem rolagem lateral */
        @media (max-width: 560px) {{
            .tabela-empilhavel thead {{ display: none; }}
            .tabela-empilhavel, .tabela-empilhavel tbody {{ display: block; width: 100%; }}
            .tabela-empilhavel tr {{
                display: block; width: 100%; padding: 10px 12px; margin-bottom: 8px;
                border: 1px solid #e9ecef; border-radius: 8px; background: white;
            }}
            .tabela-empilhavel tr[colspan], .tabela-empilhavel tr:has(td[colspan]) {{
                border: none; background: transparent; padding: 8px 0;
            }}
            .tabela-empilhavel td {{
                display: flex; justify-content: space-between; align-items: baseline;
                gap: 10px; padding: 4px 0 !important; border: none !important;
                text-align: right; white-space: normal; max-width: none;
            }}
            .tabela-empilhavel td[colspan] {{ display: block; text-align: center; }}
            .tabela-empilhavel td::before {{
                content: attr(data-label); font-weight: 700; color: #8492a6;
                font-size: 10px; text-transform: uppercase; letter-spacing: 0.4px;
                text-align: left; flex-shrink: 0;
            }}
            .tabela-empilhavel td:not([data-label])::before {{ content: none; }}
            .tabela-empilhavel .mat-total::before, .tabela-empilhavel .valor::before {{ color: #718096; }}
            .linha-fornecedor td, .linha-fornecedor-pago td {{ background: transparent !important; }}
            .linha-parcela td, .linha-parcela-pago td {{ padding-left: 14px !important; }}
            .linha-parcela .fornecedor-nome, .linha-parcela-pago .fornecedor-nome {{ padding-left: 0; }}
            .resp-bloco .table-wrapper {{ padding: 8px; }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Relatório {obra_id} - Gestão de Obras</h1>
            <div class="subtitle">{obra_nome} · {subtitulo}</div>
        </div>
        <div class="tabs-nav">
            <a class="tab-btn" href="#tab-geral">📊 Visão Geral</a>
            <a class="tab-btn" href="#tab-pagamentos">💰 Pagamentos</a>
            <a class="tab-btn" href="#tab-compras-etapa">🏗️ Compras por Etapa</a>
            <a class="tab-btn" href="#tab-tipo-gasto">🧰 Tipo de Gasto</a>
            <a class="tab-btn" href="#tab-responsavel">👤 Responsável</a>
        </div>
        <div class="content">
            <div id="tab-geral" class="tab-panel">
            <div class="section-title">FLUXO DE CAIXA - {obra_id}</div>
            <div class="cards-container">
                <div class="card">
                    <div class="card-icon">💵</div>
                    <div class="card-title">Total Pago</div>
                    <div class="card-value">{fmt_money(total_pago)}</div>
                </div>
                <div class="card">
                    <div class="card-icon">💳</div>
                    <div class="card-title">Total a Pagar</div>
                    <div class="card-value">{fmt_money(total_a_pagar)}</div>
                </div>
            </div>
            <div class="cards-container">
                <div class="card alert-card">
                    <div class="card-icon">⚠️</div>
                    <div class="card-title">Vencimentos da Semana</div>
                    <div class="card-value">{fmt_money(vencimentos_semana)}</div>
                </div>
            </div>
            <div class="chart-container">
                <div class="chart-title">Fluxo de Caixa Mensal</div>
                <div class="chart-legend">
                    <div class="legend-item"><div class="legend-color legend-pago"></div><span>Pago</span></div>
                    <div class="legend-item"><div class="legend-color legend-a-pagar"></div><span>A Pagar</span></div>
                </div>
                <div class="bars-container">{barras_html}
                </div>
            </div>
            <div class="section-title">CUSTO POR ETAPA - {obra_id}</div>
            <div class="etapas-container">{etapas_html}
            </div>
            <div class="totais-container">
                <div class="total-card acumulado">
                    <div class="total-label">Custo Total Geral Acumulado</div>
                    <div class="total-value">{fmt_money(total_geral)}</div>
                </div>
                <div class="total-card quitado">
                    <div class="total-label">Total Quitado</div>
                    <div class="total-value">{fmt_money(total_pago)}</div>
                </div>
            </div>
            <div class="a-pagar-section">
                <div class="section-title-a-pagar">
                    <span>A Pagar por Fornecedor</span>
                    <span class="total-a-pagar">{fmt_money(total_a_pagar)}</span>
                </div>
                <div class="table-wrapper" style="border:none; border-radius:0;">
                <table class="tabela-a-pagar tabela-empilhavel">
                    <thead>
                        <tr><th>Fornecedor</th><th>Vencimento</th><th>Valor</th></tr>
                    </thead>
                    <tbody>{apagar_html}
                    </tbody>
                </table>
                </div>
            </div>
            <div class="materiais-busca-wrap">
                <input type="search" id="buscaMaterial" class="materiais-busca"
                       placeholder="🔎 Filtrar por material... (ex: cimento)"
                       autocomplete="off" autocorrect="off" autocapitalize="off" spellcheck="false"
                       oninput="filtrarMateriais()">
                <span id="buscaMaterialInfo" class="materiais-busca-info"></span>
            </div>
            <noscript>
                <div class="aviso-sem-js">
                    ⚠️ A busca acima precisa de JavaScript pra funcionar e parece estar desativado
                    neste visualizador (comum ao abrir arquivos direto pelo WhatsApp/Telegram no
                    celular). Toque em Compartilhar → <strong>Abrir no Safari</strong> (ou Chrome)
                    pra usar a busca. Todos os dados abaixo continuam completos e visíveis normalmente.
                </div>
            </noscript>
            <div class="materiais-section">
                <div class="materiais-title">📦 Quantidade Total por Material — {obra_id}</div>
                <div class="materiais-subtitle">Consolidado de todas as etapas e notas lançadas</div>
                <div class="table-wrapper">
                    <table class="materiais-table tabela-empilhavel" id="tabelaMatTotal">
                        <thead>
                            <tr><th>Produto</th><th>Qtd. Total</th><th>Custo UND Médio</th><th>Custo Total</th><th>Etapa(s)</th></tr>
                        </thead>
                        <tbody>{mat_total_html}
                        </tbody>
                    </table>
                    <div class="materiais-vazio" id="vazioMatTotal" style="display:none;">Nenhum material encontrado.</div>
                </div>
            </div>
            <div class="materiais-section">
                <div class="materiais-title">📦 Detalhamento por Etapa — {obra_id}</div>
                <div class="table-wrapper">
                    <table class="materiais-table tabela-empilhavel" id="tabelaMatDetalhe">
                        <thead>
                            <tr><th>Produto</th><th>Etapa</th><th>Quantidade</th><th>Custo UND</th><th>Custo Total</th></tr>
                        </thead>
                        <tbody>{materiais_html}
                        </tbody>
                    </table>
                    <div class="materiais-vazio" id="vazioMatDetalhe" style="display:none;">Nenhum material encontrado.</div>
                </div>
            </div>
            </div>
            <div id="tab-pagamentos" class="tab-panel">
                <div class="pago-section">
                    <div class="section-title-a-pagar">
                        <span>Pagamentos Realizados por Fornecedor</span>
                        <span class="total-pago-fornecedor">{fmt_money(total_pago)}</span>
                    </div>
                    <div class="table-wrapper" style="border:none; border-radius:0;">
                    <table class="tabela-a-pagar tabela-empilhavel">
                        <thead>
                            <tr><th>Fornecedor</th><th>Data</th><th>Valor</th></tr>
                        </thead>
                        <tbody>{pago_html}
                        </tbody>
                    </table>
                    </div>
                </div>
            </div>
            <div id="tab-compras-etapa" class="tab-panel">
                <div class="section-title">TOTAL GASTO POR ETAPA - {obra_id}</div>
                <div class="etapas-container">{compras_resumo_html}
                </div>
                <div class="section-title">LANÇAMENTOS AGRUPADOS POR ETAPA - {obra_id}</div>
                {compras_detalhe_html}
            </div>
            <div id="tab-tipo-gasto" class="tab-panel">
                <div class="section-title">TOTAL POR TIPO DE GASTO - {obra_id}</div>
                <div class="materiais-subtitle" style="margin-top:-10px;">Material, mão de obra, equipamento etc — independente da etapa onde foi usado</div>
                <div class="etapas-container">{tipo_resumo_html}
                </div>
                <div class="section-title">LANÇAMENTOS AGRUPADOS POR TIPO - {obra_id}</div>
                {tipo_detalhe_html}
            </div>
            <div id="tab-responsavel" class="tab-panel">
                <div class="section-title">TOTAL PAGO/A PAGAR POR RESPONSÁVEL - {obra_id}</div>
                <div class="etapas-container">{resp_resumo_html}
                </div>
                <div class="section-title">DETALHAMENTO DAS NOTAS POR RESPONSÁVEL - {obra_id}</div>
                {resp_detalhe_html}
            </div>
            <div class="footer">Relatório gerado em {hoje.strftime("%d/%m/%Y")} às {datetime.datetime.now().strftime("%H:%M")}</div>
        </div>
    </div>
    <script>
        function filtrarMateriais() {{
            var campo = document.getElementById('buscaMaterial');
            if (!campo) return;
            var termo = campo.value.trim().toLowerCase();
            var tabelas = [
                {{tabela: 'tabelaMatTotal', vazio: 'vazioMatTotal'}},
                {{tabela: 'tabelaMatDetalhe', vazio: 'vazioMatDetalhe'}}
            ];
            var totalVisiveis = 0;
            tabelas.forEach(function(cfg) {{
                var linhas = document.querySelectorAll('#' + cfg.tabela + ' tbody tr.linha-material');
                var visiveis = 0;
                linhas.forEach(function(tr) {{
                    var produto = tr.getAttribute('data-produto') || '';
                    var mostra = termo === '' || produto.indexOf(termo) !== -1;
                    tr.style.display = mostra ? '' : 'none';
                    if (mostra) visiveis++;
                }});
                var vazioEl = document.getElementById(cfg.vazio);
                if (vazioEl) vazioEl.style.display = (linhas.length > 0 && visiveis === 0) ? 'block' : 'none';
                totalVisiveis += visiveis;
            }});
            var info = document.getElementById('buscaMaterialInfo');
            if (info) info.textContent = termo === '' ? '' : totalVisiveis + ' resultado(s)';
        }}
        // Reforço para celular: alguns teclados/navegadores móveis não disparam o
        // "oninput" inline de forma confiável, então religamos por vários eventos.
        (function() {{
            function ligar() {{
                var campo = document.getElementById('buscaMaterial');
                if (!campo) return;
                ['input', 'keyup', 'change', 'search', 'paste'].forEach(function(evento) {{
                    campo.addEventListener(evento, filtrarMateriais, false);
                }});
            }}
            if (document.readyState === 'loading') {{
                document.addEventListener('DOMContentLoaded', ligar);
            }} else {{
                ligar();
            }}
        }})();
    </script>
</body>
</html>'''
    return html


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Uso: python3 gerar_relatorio.py <planilha.xlsx> <ID_OBRA> [saida.html]")
        sys.exit(1)
    planilha = sys.argv[1]
    obra_id = sys.argv[2]
    saida = sys.argv[3] if len(sys.argv) > 3 else f"relatorio_{obra_id}.html"
    html = build_report(obra_id, planilha)
    with open(saida, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Relatório gerado: {saida}")
